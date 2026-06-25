# api.py
"""
api.py - Serveur Flask pour l'interface React
Version asynchrone avec jobs, polling HiGHS, et persistance des runs.
Correction : chemin temporaire Windows compatible + attente du fichier de log.
Ajout du support de l'arbre Branch & Bound (campagne activée).
Ajout du champ prix_vente dans commandes_acceptees_detail.
Ajout des surcharges de paramètres (overrides) et des contraintes bloquantes.
Corrections: harmonisation des stocks (arrondi, semaine 0), agrégation HRC, endpoint lecture paramètres.
"""

import os
import tempfile
import json
import time
import threading
import uuid
import re
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from data_loader1 import charger_donnees, charger_donnees_depuis
from model1 import construire_modele
from main import resoudre, analyser_resultats, exporter_resultats, parser_bnb_log, diagnostiquer_refus
import pyomo.environ as pyo
import openpyxl

app = Flask(__name__)
CORS(app)

# Paramètres par défaut
DEFAULT_B2 = True
DEFAULT_B4 = False
DEFAULT_GAP = 0.01  # 1%

EXPORT_DIR = "exports"
os.makedirs(EXPORT_DIR, exist_ok=True)

# Stockage des jobs en mémoire (volatil)
JOBS = {}
JOBS_LOCK = threading.Lock()

# Persistance des runs sauvegardés
RUNS_STORE = "saved_runs.json"


# ----------------------------------------------------------------------
# Fonctions d'aide pour le log HiGHS
# ----------------------------------------------------------------------
def parse_highs_gap(line: str):
    """Extrait le MIP gap (en %) d'une ligne de log HiGHS."""
    pcts = re.findall(r'(\d+\.\d+)%', line)
    if len(pcts) >= 2:
        try:
            return float(pcts[1])
        except ValueError:
            return None
    return None


# ----------------------------------------------------------------------
# Helper pour déterminer la raison principale de refus (TÂCHE 2.2)
# ----------------------------------------------------------------------
def _determiner_raison_principale(cmd_i: dict) -> str:
    """Détermine la raison principale de refus basée sur les attributs de la commande."""
    prio = cmd_i.get("priorite", "")
    if prio == "Basse":
        return "Priorité basse"
    elif prio == "Haute":
        return "Contrainte capacité/ressource"
    else:
        return "Capacité insuffisante"


# ----------------------------------------------------------------------
# Fonctions d'arrondi et d'agrégation
# ----------------------------------------------------------------------
def round_stock_value(val):
    """Arrondit une valeur de stock à 1 décimale."""
    return round(val, 1) if val is not None else 0.0

def aggregate_conso_hrc(m, data, commandes_acceptees):
    """Agrège la consommation HRC par grade et semaine (fonction commune)."""
    conso_by_grade_semaine = {}
    for i in commandes_acceptees:
        cmd_i = data["commandes"][i]
        grade = cmd_i["grade"]
        for p in range(len(cmd_i["chemins"])):
            if "PK" in cmd_i["chemins"][p]:
                for t in data["T"]:
                    val = pyo.value(m.x[i, p, "PK", t])
                    if val > 0.01:
                        key = (grade, t)
                        conso_by_grade_semaine[key] = conso_by_grade_semaine.get(key, 0.0) + val

    result = []
    for (grade, t), total in conso_by_grade_semaine.items():
        result.append({
            'grade': grade,
            'semaine': t,
            'stock': round_stock_value(total)  # arrondi à 1 décimale
        })
    # Ajouter les couples (grade, semaine) manquants avec stock=0
    for g in data["G"]:
        for t in data["T"]:
            if not any(r['grade'] == g and r['semaine'] == t for r in result):
                result.append({'grade': g, 'semaine': t, 'stock': 0.0})
    return result


# ----------------------------------------------------------------------
# Fonction d'exécution du job (dans un thread séparé)
# ----------------------------------------------------------------------
def run_optimization_job(job_id, temp_path, activer_B2, activer_B4, mip_gap, overrides):
    try:
        # 1. Charger les données
        import data_loader1
        data_loader1.EXCEL_PATH = temp_path
        data = charger_donnees()

        # ── APPLICATION DES SURCHARGES (TÂCHE 1.1) ─────────────────────────────
        def _parse_float(val):
            """Convertit une string en float, retourne None si vide ou invalide."""
            if val is None or str(val).strip() == "":
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        # 1. Paramètres économiques scalaires
        champs_params = [
            ("prix_zinc",        "prix_zinc"),
            ("conso_zinc_hdg",   "conso_zinc_hdg"),
            ("conso_zinc_ppgi",  "conso_zinc_ppgi"),
            ("prix_peinture",    "prix_peinture"),
            ("conso_peinture",   "conso_peinture"),
            ("pen_haute",        "pen_haute"),
            ("pen_normale",      "pen_normale"),
            ("pen_basse",        "pen_basse"),
            ("cout_stock_inter", "cout_stock_inter"),
            ("cout_stock_fini",  "cout_stock_fini"),
            ("prix_chute",       "prix_chute"),
            ("coef_decl",        "coef_decl"),
            ("coef_nc",          "coef_nc"),
        ]
        for json_key, data_key in champs_params:
            val = _parse_float(overrides.get(json_key, ""))
            if val is not None:
                data["params"][data_key] = val
                print(f"[OVERRIDE] params[{data_key}] = {val}")

        # 2. Stocks initiaux PK (par grade)
        for grade, stock_override in overrides.get("stock_pk", {}).items():
            if grade in data["stock_pk"] and isinstance(stock_override, dict):
                for champ in ("init", "min", "max"):
                    v = stock_override.get(champ)
                    if v is not None:
                        data["stock_pk"][grade][champ] = float(v)
                        print(f"[OVERRIDE] stock_pk[{grade}][{champ}] = {float(v)}")

        # 3. Stocks interprocess
        for point, stock_override in overrides.get("stock_inter", {}).items():
            if point in data["stock_inter"] and isinstance(stock_override, dict):
                for champ in ("init", "min", "max"):
                    v = stock_override.get(champ)
                    if v is not None:
                        data["stock_inter"][point][champ] = float(v)
                        print(f"[OVERRIDE] stock_inter[{point}][{champ}] = {float(v)}")

        # 4. Stocks produits finis
        for famille, stock_override in overrides.get("stock_fini", {}).items():
            if famille in data["stock_fini"] and isinstance(stock_override, dict):
                for champ in ("init", "min", "max"):
                    v = stock_override.get(champ)
                    if v is not None:
                        data["stock_fini"][famille][champ] = float(v)
                        print(f"[OVERRIDE] stock_fini[{famille}][{champ}] = {float(v)}")

        # 5. Disponibilité HRC
        for grade, val_str in overrides.get("dispo_hrc", {}).items():
            val = _parse_float(val_str)
            if val is not None and grade in data["dispo_hrc"]:
                data["dispo_hrc"][grade] = val
                print(f"[OVERRIDE] dispo_hrc[{grade}] = {val}")

        # 6. Arrêts planifiés
        # Clé format "MACHINE-St" ex: "PK-S1", "CRMA-S2"
        for key, val_str in overrides.get("arrets", {}).items():
            val = _parse_float(val_str)
            if val is not None:
                try:
                    # Le format est MACHINE-St, machine peut contenir des tirets (ex: FH-CRMA)
                    # On split sur le dernier "-S"
                    last_s_idx = key.rfind("-S")
                    if last_s_idx == -1:
                        continue
                    machine = key[:last_s_idx]
                    semaine = int(key[last_s_idx + 2:])
                    if machine in data["M"] and semaine in data["T"]:
                        data["arrets"][(machine, semaine)] = val
                        print(f"[OVERRIDE] arrets[({machine}, {semaine})] = {val}")
                except (ValueError, IndexError):
                    pass
        # ── FIN DES SURCHARGES ─────────────────────────────────────

        # 2. Construire le modèle
        m, IP, IPM = construire_modele(data, activer_B2=activer_B2, activer_B4=activer_B4)

        # 3. Préparer le fichier de log (chemin temporaire Windows compatible)
        log_dir = tempfile.gettempdir()
        log_path = os.path.join(log_dir, f"{job_id}_solver.log")

        # Lancer le thread qui tail le log (avec attente de création du fichier)
        def tail_log():
            # Attendre que le fichier de log existe (max 10 secondes)
            waited = 0
            while not os.path.exists(log_path) and waited < 10:
                time.sleep(0.5)
                waited += 0.5
            if not os.path.exists(log_path):
                # Si le fichier n'existe toujours pas, on ne peut pas tailer
                return

            with open(log_path, 'r') as f:
                # Se placer à la fin pour ne lire que les nouvelles lignes
                f.seek(0, os.SEEK_END)
                while True:
                    line = f.readline()
                    if not line:
                        time.sleep(0.2)
                        continue
                    gap = parse_highs_gap(line)
                    if gap is not None:
                        with JOBS_LOCK:
                            JOBS[job_id]["gap"] = gap
                    # Vérifier si le job est terminé
                    with JOBS_LOCK:
                        if JOBS[job_id]["status"] in ("done", "error"):
                            break

        threading.Thread(target=tail_log, daemon=True).start()

        # 4. Résoudre avec capture du log
        start_time = time.time()
        results = resoudre(m, time_limit=600, mip_gap=mip_gap, log_file=log_path)
        wall_time = time.time() - start_time

        # 5. Analyser les résultats
        refusees, retards, goulots = analyser_resultats(m, data, IP, IPM)
        commandes_acceptees = [i for i in data["I"] if pyo.value(m.y[i]) > 0.5]
        commandes_refusees = [i for i in data["I"] if pyo.value(m.y[i]) < 0.5]

        # 6. Construire le dictionnaire de résultats
        result_data = {
            'statut': str(results.termination_condition),
            'marge': pyo.value(m.obj),
            'temps_execution': round(wall_time, 2),
            'commandes_acceptees': commandes_acceptees,
            # TÂCHE 2.2 - enrichissement des commandes refusées avec contraintes bloquantes
            'commandes_refusees': [
                {
                    'id': i,
                    'famille': data["commandes"][i]["famille"],
                    'grade': data["commandes"][i]["grade"],
                    'tonnage': data["commandes"][i]["tonnage"],
                    'prix': data["commandes"][i]["prix"],
                    'priorite': data["commandes"][i]["priorite"],
                    'semaine_liv': data["commandes"][i]["semaine_liv"],
                    'client': data["commandes"][i]["client"],
                    'raison_principale': _determiner_raison_principale(data["commandes"][i]),
                    'contraintes_bloquantes': diagnostiquer_refus(m, data, i),
                }
                for i in commandes_refusees
            ],
            'retards': [
                {
                    'id': r[0],
                    'famille': r[1],
                    'tonnage': r[2],
                    'semaine_prevue': r[3],
                    'semaine_produite': r[4],
                    'retard_semaines': r[5]
                }
                for r in retards
            ],
            'goulots': [
                {'machine': g[0], 'semaine': g[1], 'taux_utilisation': g[2]}
                for g in goulots
            ],
            'gap': mip_gap,
        }

        # --- Enrichissements (inchangés) ---
        prix_hrc = data["prix_hrc"]
        rendements = data["rendements"]
        cmd = data["commandes"]

        acceptees_detail = []
        for i in commandes_acceptees:
            cmd_i = cmd[i]
            f = cmd_i["famille"]
            tonnage = cmd_i["tonnage"]
            prix_vente = cmd_i["prix"]
            chemin_typique = cmd_i["chemins"][0]
            rendement_total = 1.0
            for mach in chemin_typique:
                rendement_total *= rendements[mach]
            prix_hrc_cmd = prix_hrc.get((cmd_i["grade"], cmd_i["largeur"]), 6000)
            cout_hrc_par_tonne = prix_hrc_cmd / rendement_total
            marge_unitaire = prix_vente - cout_hrc_par_tonne
            marge_totale = marge_unitaire * tonnage

            sem_prod = None
            for p in range(len(cmd_i["chemins"])):
                last_mach = cmd_i["chemins"][p][-1]
                for t in data["T"]:
                    if pyo.value(m.x[i, p, last_mach, t]) > 0.01:
                        sem_prod = t
                        break
                if sem_prod:
                    break

            acceptees_detail.append({
                "id": i,
                "client": cmd_i["client"],
                "famille": f,
                "grade": cmd_i["grade"],
                "epaisseur": cmd_i["epaisseur"],
                "largeur": cmd_i["largeur"],
                "tonnage": tonnage,
                "prix_vente": prix_vente,
                "semaine_liv": cmd_i["semaine_liv"],
                "priorite": cmd_i["priorite"],
                "semaine_prod": sem_prod,
                "marge_mad": round(marge_totale, 2),
                "marge_unitaire_mad": round(marge_unitaire, 2),
                "statut": "Acceptée"
            })
        result_data['commandes_acceptees_detail'] = acceptees_detail

        # Production par semaine et famille
        prod_sem = {f: {t: 0 for t in data["T"]} for f in data["F"]}
        for i in commandes_acceptees:
            cmd_i = cmd[i]
            for p in range(len(cmd_i["chemins"])):
                last_mach = cmd_i["chemins"][p][-1]
                rho = rendements[last_mach]
                for t in data["T"]:
                    val = pyo.value(m.x[i, p, last_mach, t])
                    if val > 0.01:
                        prod_sem[cmd_i["famille"]][t] += rho * val
        result_data['production_par_semaine'] = [
            {
                'semaine': t,
                'CRC': prod_sem['CRC'][t],
                'HDG': prod_sem['HDG'][t],
                'PPGI': prod_sem['PPGI'][t],
                'BACR': prod_sem['BACR'][t]
            }
            for t in data["T"]
        ]

        # Marge par famille
        marge_fam = {f: 0.0 for f in data["F"]}
        for i in commandes_acceptees:
            cmd_i = cmd[i]
            f = cmd_i["famille"]
            chemin_typique = cmd_i["chemins"][0]
            rendement_total = 1.0
            for mach in chemin_typique:
                rendement_total *= rendements[mach]
            prix_hrc_cmd = prix_hrc.get((cmd_i["grade"], cmd_i["largeur"]), 6000)
            cout_hrc_par_tonne = prix_hrc_cmd / rendement_total
            marge_unitaire = cmd_i["prix"] - cout_hrc_par_tonne
            marge_fam[f] += marge_unitaire * cmd_i["tonnage"]
        total_marge = sum(marge_fam.values())
        result_data['marge_par_famille'] = [
            {'name': f, 'value': marge_fam[f], 'pct': round(100 * marge_fam[f] / total_marge, 1) if total_marge > 0 else 0}
            for f in data["F"] if marge_fam[f] > 0
        ]

        # Statut des commandes
        statut_counts = {
            'Honorées à l\'échéance': 0,
            'Honorées en avance': 0,
            'Honorées en retard': 0,
            'Refusées': 0
        }
        for i in commandes_acceptees:
            cmd_i = cmd[i]
            sem_prod = None
            for p in range(len(cmd_i["chemins"])):
                last_mach = cmd_i["chemins"][p][-1]
                for t in data["T"]:
                    if pyo.value(m.x[i, p, last_mach, t]) > 0.01:
                        sem_prod = t
                        break
                if sem_prod:
                    break
            if sem_prod is None:
                continue
            if sem_prod < cmd_i["semaine_liv"]:
                statut_counts['Honorées en avance'] += cmd_i["tonnage"]
            elif sem_prod == cmd_i["semaine_liv"]:
                statut_counts['Honorées à l\'échéance'] += cmd_i["tonnage"]
            else:
                statut_counts['Honorées en retard'] += cmd_i["tonnage"]
        for ref in result_data['commandes_refusees']:
            statut_counts['Refusées'] += ref['tonnage']
        total_tonnage = sum(statut_counts.values())
        result_data['statut_commandes'] = [
            {'name': k, 'value': v, 'pct': round(100 * v / total_tonnage, 1) if total_tonnage > 0 else 0}
            for k, v in statut_counts.items()
        ]

        # Utilisation des lignes (avec capacité)
        cadences = data["cadences"]
        arrets = data["arrets"]
        jours = data["params"]["jours_semaine"]
        utilisation_data = []
        for mach in data["M"]:
            for t in data["T"]:
                utilis = 0
                capa_total = 0
                for f in data["F"]:
                    cad = cadences.get((mach, f), 0)
                    if cad > 0:
                        capa_total += cad * (jours - arrets.get((mach, t), 0))
                        for i in commandes_acceptees:
                            cmd_i = cmd[i]
                            if cmd_i["famille"] != f:
                                continue
                            for p in range(len(cmd_i["chemins"])):
                                if mach in cmd_i["chemins"][p]:
                                    utilis += pyo.value(m.x[i, p, mach, t])
                if capa_total > 0:
                    utilisation_data.append({
                        'machine': mach,
                        'semaine': t,
                        'tonnage': round(utilis, 1),
                        'capacite': round(capa_total, 1),
                        'taux': round(100 * utilis / capa_total, 1)
                    })
        result_data['utilisation_lignes'] = utilisation_data

        # Stocks (PK, inter, finis) - avec arrondi et semaines 0..4
        pk_stocks = []
        for g in data["G"]:
            for t in [0] + data["T"]:
                if hasattr(m.stockPK[g, t], 'value'):
                    pk_stocks.append({'grade': g, 'semaine': t, 'stock': round_stock_value(pyo.value(m.stockPK[g, t]))})
        result_data['stocks'] = {'pk': pk_stocks}
        inter_stocks = []
        for k in data["K"]:
            for t in [0] + data["T"]:
                if hasattr(m.Iinter[k, t], 'value'):
                    inter_stocks.append({'point': k, 'semaine': t, 'stock': round_stock_value(pyo.value(m.Iinter[k, t]))})
        result_data['stocks']['inter'] = inter_stocks
        fini_stocks = []
        for f in data["F"]:
            for t in [0] + data["T"]:
                if hasattr(m.StockPhysique[f, t], 'value'):
                    fini_stocks.append({'famille': f, 'semaine': t, 'stock': round_stock_value(pyo.value(m.StockPhysique[f, t]))})
        result_data['stocks']['fini'] = fini_stocks

        # --- NOUVEAU : Consommation HRC par grade et semaine (agrégée) ---
        result_data['conso_hrc_semaine'] = aggregate_conso_hrc(m, data, commandes_acceptees)

        # StockChartData
        stock_chart = []
        for t in [0] + data["T"]:
            row = {"t": f"t={t}"}
            row["PK"] = sum([s["stock"] for s in pk_stocks if s["semaine"] == t])
            row["Interprocess"] = sum([s["stock"] for s in inter_stocks if s["semaine"] == t])
            row["Finis HDG"] = sum([s["stock"] for s in fini_stocks if s["famille"] == "HDG" and s["semaine"] == t])
            row["Finis CRC"] = sum([s["stock"] for s in fini_stocks if s["famille"] == "CRC" and s["semaine"] == t])
            stock_chart.append(row)
        result_data['stockChartData'] = stock_chart

        # Plan de production détaillé
        plan_production = []
        for i in commandes_acceptees:
            cmd_i = cmd[i]
            for p in range(len(cmd_i["chemins"])):
                for mach in cmd_i["chemins"][p]:
                    for t in data["T"]:
                        val = pyo.value(m.x[i, p, mach, t])
                        if val > 0.01:
                            plan_production.append({
                                'commande': i,
                                'famille': cmd_i["famille"],
                                'grade': cmd_i["grade"],
                                'chemin': p,
                                'machine': mach,
                                'semaine': t,
                                'tonnage_entrant': round(val, 1),
                                'rendement': data["rendements"][mach],
                                'tonnage_sortant': round(val * data["rendements"][mach], 1),
                            })
        result_data['plan_production'] = plan_production

        # Export Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_filename = f"resultats_maghreb_steel_{timestamp}.xlsx"
        export_path = os.path.join(EXPORT_DIR, export_filename)
        exporter_resultats(
            m, data,
            commandes_acceptees,
            refusees,
            retards,
            goulots,
            wall_time,
            filepath=export_path
        )
        result_data['export_file'] = export_filename

        # --- TÂCHE 3 : Parser l'arbre B&B si campagne activée ---
        # Mettre le statut à "done" pour arrêter le thread tail_log
        with JOBS_LOCK:
            JOBS[job_id]["status"] = "done"
            JOBS[job_id]["result"] = result_data

        # Maintenant parser le B&B (fichier libéré)
        bnb_data = None
        if activer_B4 and os.path.exists(log_path):
            bnb_data = parser_bnb_log(log_path)
            if bnb_data:
                result_data['bnb_tree'] = bnb_data
                with JOBS_LOCK:
                    JOBS[job_id]["result"] = result_data
            else:
                print(f"[INFO] Pas de données B&B dans {log_path}")
        else:
            print(f"[INFO] Pas de parsing B&B (activer_B4={activer_B4}, log_path exist={os.path.exists(log_path)})")

        # ── Test de cohérence des stocks (Bug 1) ──────────────────────────
        try:
            wb = openpyxl.load_workbook(export_path)
            if "Stocks_Physiques" in wb.sheetnames:
                ws = wb["Stocks_Physiques"]
                # Somme des stocks physiques sur toutes les lignes (colonnes C)
                sum_excel = 0.0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[2] is not None:
                        sum_excel += row[2]
                sum_json = sum(s["stock"] for s in result_data['stocks']['fini'])
                if abs(sum_excel - sum_json) > 0.001 * max(1.0, sum_json):
                    print(f"[WARNING] Incohérence stocks finis: Excel={sum_excel:.1f}, JSON={sum_json:.1f}")
                else:
                    print(f"[OK] Cohérence stocks finis: Excel={sum_excel:.1f}, JSON={sum_json:.1f}")
        except Exception as e:
            print(f"[WARNING] Impossible de vérifier la cohérence des stocks: {e}")

        # Nettoyer le fichier temporaire
        os.unlink(temp_path)

    except Exception as e:
        with JOBS_LOCK:
            JOBS[job_id]["status"] = "error"
            JOBS[job_id]["error"] = str(e)
        if os.path.exists(temp_path):
            os.unlink(temp_path)


# ----------------------------------------------------------------------
# Routes Flask
# ----------------------------------------------------------------------
@app.route('/optimize', methods=['POST'])
def optimize_start():
    """Démarre un job d'optimisation de manière asynchrone."""
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier envoyé'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nom de fichier vide'}), 400

    activer_B2 = request.form.get('activer_B2', str(DEFAULT_B2)).lower() == 'true'
    activer_B4 = request.form.get('activer_B4', str(DEFAULT_B4)).lower() == 'true'
    mip_gap = float(request.form.get('gap', DEFAULT_GAP)) / 100

    # TÂCHE 1.2 - Lire les overrides depuis le FormData
    overrides_raw = request.form.get('overrides', '{}')
    try:
        overrides = json.loads(overrides_raw)
    except (json.JSONDecodeError, TypeError):
        overrides = {}

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        file.save(tmp.name)
        temp_path = tmp.name

    job_id = str(uuid.uuid4())
    with JOBS_LOCK:
        JOBS[job_id] = {"status": "running", "gap": None, "result": None, "error": None}

    # TÂCHE 1.2 - Passer overrides dans le thread
    thread = threading.Thread(
        target=run_optimization_job,
        args=(job_id, temp_path, activer_B2, activer_B4, mip_gap, overrides),
        daemon=True
    )
    thread.start()

    return jsonify({"job_id": job_id})


@app.route('/optimize/status/<job_id>')
def optimize_status(job_id):
    """Retourne l'état du job (status, gap, résultat ou erreur)."""
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "job inconnu"}), 404
    response = {
        "status": job["status"],
        "gap": job.get("gap"),
        "error": job.get("error"),
    }
    if job["status"] == "done" and job.get("result"):
        response["result"] = job["result"]
    return jsonify(response)


# ----------------------------------------------------------------------
# Nouvel endpoint : lecture des paramètres d'un fichier Excel
# ----------------------------------------------------------------------
@app.route('/lire-parametres-fichier', methods=['POST'])
def lire_parametres_fichier():
    """Reçoit un fichier Excel, le charge et renvoie les paramètres de surcharge."""
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier envoyé'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nom de fichier vide'}), 400

    try:
        # Sauvegarder le fichier temporairement
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            temp_path = tmp.name

        # Charger les données depuis ce fichier
        data = charger_donnees_depuis(temp_path)

        # Extraire les valeurs pour les overrides
        params = data["params"]
        overrides = {
            "prix_zinc": str(params["prix_zinc"]),
            "conso_zinc_hdg": str(params["conso_zinc_hdg"]),
            "conso_zinc_ppgi": str(params["conso_zinc_ppgi"]),
            "prix_peinture": str(params["prix_peinture"]),
            "conso_peinture": str(params["conso_peinture"]),
            "pen_haute": str(params["pen_haute"]),
            "pen_normale": str(params["pen_normale"]),
            "pen_basse": str(params["pen_basse"]),
            "cout_stock_inter": str(params["cout_stock_inter"]),
            "cout_stock_fini": str(params["cout_stock_fini"]),
            "prix_chute": str(params["prix_chute"]),
            "coef_decl": str(params["coef_decl"]),
            "coef_nc": str(params["coef_nc"]),
            "stock_pk": {},
            "stock_inter": {},
            "stock_fini": {},
            "dispo_hrc": {},
            "arrets": {},
        }

        # Stocks PK
        for grade, vals in data["stock_pk"].items():
            overrides["stock_pk"][grade] = {
                "init": vals["init"],
                "min": vals["min"],
                "max": vals["max"],
            }

        # Stocks interprocess
        for point, vals in data["stock_inter"].items():
            overrides["stock_inter"][point] = {
                "init": vals["init"],
                "min": vals["min"],
                "max": vals["max"],
            }

        # Stocks finis
        for famille, vals in data["stock_fini"].items():
            overrides["stock_fini"][famille] = {
                "init": vals["init"],
                "min": vals["min"],
                "max": vals["max"],
            }

        # Disponibilité HRC
        for grade, dispo in data["dispo_hrc"].items():
            overrides["dispo_hrc"][grade] = str(dispo)

        # Arrêts planifiés (format clé "MACHINE-S1")
        for (machine, semaine), jours in data["arrets"].items():
            key = f"{machine}-S{semaine}"
            overrides["arrets"][key] = str(jours)

        os.unlink(temp_path)
        return jsonify(overrides)

    except Exception as e:
        return jsonify({'error': f'Erreur de lecture du fichier: {str(e)}'}), 400


# ----------------------------------------------------------------------
# Persistance des runs
# ----------------------------------------------------------------------
@app.route('/runs', methods=['POST'])
def save_run():
    run_data = request.json
    runs = []
    if os.path.exists(RUNS_STORE):
        with open(RUNS_STORE, 'r', encoding='utf-8') as f:
            try:
                runs = json.load(f)
            except json.JSONDecodeError:
                runs = []
    runs.append(run_data)
    with open(RUNS_STORE, 'w', encoding='utf-8') as f:
        json.dump(runs, f, ensure_ascii=False, indent=2)
    return jsonify({"ok": True})

@app.route('/runs/<int:run_id>', methods=['DELETE'])
def delete_run(run_id):
    """Supprime un run sauvegardé par son ID."""
    if not os.path.exists(RUNS_STORE):
        return jsonify({"error": "Aucun run sauvegardé"}), 404
    try:
        with open(RUNS_STORE, 'r', encoding='utf-8') as f:
            runs = json.load(f)
    except json.JSONDecodeError:
        runs = []
    
    initial_len = len(runs)
    # Forcer la comparaison en int pour éviter les problèmes de type
    runs = [r for r in runs if int(r.get('id', -1)) != run_id]
    
    if len(runs) == initial_len:
        return jsonify({"error": "Run non trouvé"}), 404
    
    with open(RUNS_STORE, 'w', encoding='utf-8') as f:
        json.dump(runs, f, ensure_ascii=False, indent=2)
    
    return jsonify({"ok": True})

@app.route('/runs', methods=['GET'])
def list_runs():
    if not os.path.exists(RUNS_STORE):
        return jsonify([])
    with open(RUNS_STORE, 'r', encoding='utf-8') as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError:
            data = []
    return jsonify(data)


@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(EXPORT_DIR, filename, as_attachment=True)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)